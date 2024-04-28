import codecs
import chardet
# import pandas as pd
from logger import Logger
from logger import LogLevel
import openpyxl

class TextReplacer:
    def __init__(self):
        self.rule_list = {}
        self.logger = Logger()
        self.logger.log(LogLevel.INFO, "初始化文本替换器")

    def update_files(self, input_file: str, output_file: str, rule_list_file: str):
        if not input_file:
            raise ValueError("输入文件路径不能为空。")
        else:
            self.input_file = input_file
        if not output_file:
            raise ValueError("输出文件路径不能为空。")
        else:
            self.output_file = output_file
        if not rule_list_file:
            raise ValueError("规则列表文件路径不能为空。")
        else:
            self.rule_list_file = rule_list_file
    #
    def detect_encoding(self):
        try:
            with open(self.input_file, 'rb') as f:
                result = chardet.detect(f.read())
                return result['encoding']
        except Exception as e:
            raise ValueError(f"解析{self.input_file}文件编码失败：{str(e)}")

    # def load_rule_listx(self):
    #     try:
    #         df = pd.read_excel(self.rule_list_file)
    #         for _, row in df.iterrows():
    #             old_word = str(row['原词'])
    #             new_word = str(row['替换词'])
    #             self.rule_list[old_word] = new_word
    #     except Exception as e:
    #         raise ValueError(f"解析规则列表文件失败：{str(e)}")

    def load_rule_list(self):
        try:
            wb = openpyxl.load_workbook(self.rule_list_file)
            ws = wb.active

            column_names = {cell.value: cell.column_letter for cell in ws[1]}
            old_word_column = column_names.get('原词')
            new_word_column = column_names.get('替换词')

            if old_word_column is None or new_word_column is None:
                raise ValueError("规则列表文件缺少必要的列")

            for row in ws.iter_rows(min_row=2, values_only=True):
                old_word = str(row[ord(old_word_column) - 65])  # Convert column letter to index
                new_word = str(row[ord(new_word_column) - 65])  # Convert column letter to index
                self.rule_list[old_word] = new_word
        except Exception as e:
            raise ValueError(f"解析规则列表文件失败：{str(e)}")
        finally:
            if wb:
                wb.close()

    def replace_words(self):
        try:
            encoding = self.detect_encoding()
        except Exception as e:
            encoding = 'utf-8'
        try:
            with codecs.open(self.output_file, 'w', encoding=encoding, errors='ignore') as output:
                with codecs.open(self.input_file, 'r', encoding=encoding, errors='replace') as input:
                    content = input.read()
                    for old_word, new_word in self.rule_list.items():
                        content = content.replace(old_word, new_word)
                    output.write(content)
        except Exception as e:
            self.logger.log(LogLevel.ERROR, f"替换失败：{str(e)}")
            raise ValueError(f"替换失败：{str(e)}")

# 写一个main函数获取三个文件路径，然后调用TextReplacer类的方法进行替换
# def main():
#     if len(sys.argv) != 4:
#         print("Usage: python replacer.py <input_file> <output_file> <rule_list_file>")
#         return

#     input_file = sys.argv[1]
#     output_file = sys.argv[2]
#     rule_list_file = sys.argv[3]

#     replacer = TextReplacer(input_file, output_file, rule_list_file)
#     replacer.load_rule_list()
#     replacer.replace_words()

# if __name__ == "__main__":
#     main()