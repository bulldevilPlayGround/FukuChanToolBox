import openpyxl
import re
import string
# Regular expression patterns
#时间戳可能没有小时，只有分钟，而且以小时或分钟开头时，可能只有一位数字
TIMESTAMP_PATTERN = r'^((?:\d{1,2}[:：\.])?\d{1,2}[:：\.]\d{2})\s*([\S\s]*)'
START_TEXT = "素材段数"
# videoInfo是一个视频信息的字典，key为视频组别，value包含两个列表
# 第一个列表时视频起始时间戳列表，第二个列表是视频文本列表
# 比如：
# videoInfo = {
#     "1": [
#         ["00:00", "04:23", ...],
#         ["line1", "line2", ...]
#     ],
#     "2": ...
# }
class videoInfoParserExcel:
    def __init__(self, file, verbose=False):
        self.verbose = verbose
        self.videoInfo = {}
        self.current_group = None
        self.sheet = self.__load_excel_file__(file)
        self.__fetch_video_info__(self.sheet)

    def __load_excel_file__(self, file):
        try:
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active
            if self.verbose:
                print(f"Successfully loaded {file}")
            return sheet
        except Exception as e:
            if self.verbose:
                print(f"Failed to load {file}: {e}")
            return None

    def __refine_timestamp__(self, timestamp: str) -> str:
        timestamp = re.sub(r'[.,]', ':', timestamp)

        # 分割时间戳
        parts = timestamp.split(':')
        # 补全小时、分钟、秒钟
        if len(parts) == 1:
            hours = '00'
            minutes = '00'
            seconds = parts[0].zfill(2)
        elif len(parts) == 2:
            hours = '00'
            minutes = parts[0].zfill(2)
            seconds = parts[1].zfill(2)
        else:
            hours = parts[0].zfill(2)
            minutes = parts[1].zfill(2)
            seconds = parts[2].zfill(2)
        return f"{hours}:{minutes}:{seconds}"

    def __fetch_video_info__(self, sheet):
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
        #如果第一列不为空，且为数字时，视频信息开始，这个不为空的内容为视频组别名称
            if row[0] is not None and str(row[0]).isdigit():
                # 把视频组别名称作为key，初始化视频信息字典
                print(f"视频组 {row[0]}")
                self.current_group = row[0]
                self.videoInfo[self.current_group] = [[], []]
            if self.current_group is None:
                continue

        #如果第二列不为空，且为"素材段数"时，视频信息开始，这个不为空的内容为视频组别名称
            if row[1] is not None:
                #row[1]为视频起始时间，保存到videoInfo的第一个列表中
                self.videoInfo[self.current_group][0].append(self.__refine_timestamp__(str(row[1])))

            if row[4] is not None:
                #row[4]为视频文本，保存到videoInfo的第二个列表中
                self.videoInfo[self.current_group][1].append(row[4])

    def __debug_print_video_info__(self):
        for group, info in self.videoInfo.items():
            print(f"Group: {group}, size: {len(info[0])}")
            #如果Timestamps和Texts长度不一致，打印警告
            if len(info[0]) != len(info[1]):
                print("Warning: Timestamps and Texts length mismatch")
            else:
                for index, (timestamp, text) in enumerate(zip(info[0], info[1])):
                    print(f"{index + 1}. {timestamp}: {text}")

#test
# debugMode = True
# input_file= '2.19时间线.xlsx'
# file = videoInfoParserExcel(verbose=debugMode, file=input_file)
# print(f"len of self.videoInfo: {len(file.videoInfo)}")
# file.__debug_print_video_info__()